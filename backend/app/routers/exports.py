"""
exports.py — Synchronous CSV / XLSX export router.

Schema used:
  HistoricalPrice : market, region, datetime_block, mcp_rs_mwh,
                    cleared_buy_mw, cleared_sell_mw
  Forecast        : forecast_run_id (FK), market, region, datetime_block,
                    predicted_price, lower_ci, upper_ci, confidence_level
  ForecastRun     : id (PK), forecast_date, market, region, ...
  ExportJob       : id, user_id, export_type, status, file_path, created_at
"""

import io
import uuid
from datetime import date, datetime, time
from typing import Literal, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, field_validator, model_validator
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.forecast import Forecast, ForecastRun
from app.models.jobs import ExportJob
from app.models.market import HistoricalPrice

router = APIRouter(prefix="/api/export", tags=["export"])

VALID_MARKETS = ["DAM", "GDAM", "RTM"]
FORECAST_ONLY_MARKETS = ["GDAM", "DAM", "RTM"]
MAX_DATE_RANGE_DAYS = 30


# ══════════════════════════════════════════════════════
#  datetime_block helpers
# ══════════════════════════════════════════════════════

def _dt_to_block_no(dt: datetime) -> int:
    """'2026-05-31 01:30' → block 7  (1-indexed, 15-min blocks)."""
    return (dt.hour * 60 + dt.minute) // 15 + 1


def _dt_to_time_str(dt: datetime) -> str:
    """'2026-05-31 01:30' → '01:30-01:45'."""
    start_min = dt.hour * 60 + dt.minute
    end_min = start_min + 15
    e_h, e_m = divmod(end_min % (24 * 60), 60)
    return f"{dt.hour:02d}:{dt.minute:02d}-{e_h:02d}:{e_m:02d}"


def _date_to_dt_range(d: date) -> tuple[datetime, datetime]:
    """Full day range for a single date (00:00:00 → 23:59:59)."""
    return datetime.combine(d, time.min), datetime.combine(d, time(23, 59, 59))


def _range_to_dt_range(start: date, end: date) -> tuple[datetime, datetime]:
    return datetime.combine(start, time.min), datetime.combine(end, time(23, 59, 59))


def _safe_round(val, ndigits: int = 2):
    return round(float(val), ndigits) if val is not None else ""


# ══════════════════════════════════════════════════════
#  Request Schema
# ══════════════════════════════════════════════════════

class ExportRequest(BaseModel):
    export_type: Literal["forecast", "historical", "audit"]
    market: str
    region: str = "Telangana"
    export_date: Optional[date] = None       # forecast: single date
    start_date: Optional[date] = None        # historical / audit: range start
    end_date: Optional[date] = None          # historical / audit: range end
    include_ci: bool = True

    @field_validator("market")
    @classmethod
    def validate_market(cls, v: str) -> str:
        if v not in VALID_MARKETS:
            raise ValueError(f"market must be one of {VALID_MARKETS}")
        return v

    @model_validator(mode="after")
    def validate_dates_and_market(self) -> "ExportRequest":
        if self.export_type == "forecast":
            if not self.export_date:
                raise ValueError("export_date is required for forecast export")
            if self.market not in FORECAST_ONLY_MARKETS:
                raise ValueError(
                    f"Forecast export only available for {FORECAST_ONLY_MARKETS}"
                )

        elif self.export_type == "audit":
            if self.market not in FORECAST_ONLY_MARKETS:
                raise ValueError(
                    f"Audit export only available for {FORECAST_ONLY_MARKETS} "
                    "(requires forecast data)"
                )
            self._check_range()

        else:  # historical
            self._check_range()

        return self

    def _check_range(self):
        if not self.start_date or not self.end_date:
            raise ValueError("start_date and end_date are required")
        if self.start_date > self.end_date:
            raise ValueError("start_date must be ≤ end_date")
        if (self.end_date - self.start_date).days >= MAX_DATE_RANGE_DAYS:
            raise ValueError(f"Date range cannot exceed {MAX_DATE_RANGE_DAYS} days")


# ══════════════════════════════════════════════════════
#  DB Fetchers
# ══════════════════════════════════════════════════════

async def _fetch_forecast_rows(req: ExportRequest, db: AsyncSession) -> list[dict]:
    """
    Join Forecast → ForecastRun.
    Filter by ForecastRun.forecast_date (not Forecast.forecast_date — doesn't exist).
    Use Forecast.predicted_price (not predicted_mcp).
    """
    start_dt, end_dt = _date_to_dt_range(req.export_date)

    stmt = (
        select(Forecast)
        .join(ForecastRun, Forecast.forecast_run_id == ForecastRun.id)
        .where(
            and_(
                Forecast.market == req.market,
                Forecast.region == req.region,
                Forecast.datetime_block >= start_dt,
                Forecast.datetime_block <= end_dt,
            )
        )
        .order_by(Forecast.datetime_block)
    )
    forecasts = (await db.execute(stmt)).scalars().all()

    if not forecasts:
        raise HTTPException(
            status_code=404,
            detail=(
                f"No forecast found for {req.market}/{req.region} on {req.export_date}. "
                "Run POST /api/refresh to trigger the ML pipeline."
            ),
        )

    rows = []
    for f in forecasts:
        row: dict = {
            "date": str(f.datetime_block.date()),
            "block_no": _dt_to_block_no(f.datetime_block),
            "block_time": _dt_to_time_str(f.datetime_block),
            "predicted_mcp_rs_mwh": _safe_round(f.predicted_price),
        }
        if req.include_ci:
            row["lower_ci_p10_rs_mwh"] = _safe_round(f.lower_ci)
            row["upper_ci_p90_rs_mwh"] = _safe_round(f.upper_ci)
        rows.append(row)
    return rows


async def _fetch_historical_rows(req: ExportRequest, db: AsyncSession) -> list[dict]:
    """
    Filter HistoricalPrice by datetime_block range.
    No price_date or block_no columns — both derived from datetime_block.
    """
    start_dt, end_dt = _range_to_dt_range(req.start_date, req.end_date)

    stmt = (
        select(HistoricalPrice)
        .where(
            and_(
                HistoricalPrice.market == req.market,
                HistoricalPrice.region == req.region,
                HistoricalPrice.datetime_block >= start_dt,
                HistoricalPrice.datetime_block <= end_dt,
            )
        )
        .order_by(HistoricalPrice.datetime_block)
    )
    prices = (await db.execute(stmt)).scalars().all()

    if not prices:
        raise HTTPException(
            status_code=404,
            detail=(
                f"No historical data for {req.market}/{req.region} "
                f"between {req.start_date} and {req.end_date}."
            ),
        )

    rows = []
    for p in prices:
        rows.append({
            "date": str(p.datetime_block.date()),
            "block_no": _dt_to_block_no(p.datetime_block),
            "block_time": _dt_to_time_str(p.datetime_block),
            "actual_mcp_rs_mwh": _safe_round(p.mcp_rs_mwh),
            "cleared_buy_mw": _safe_round(p.cleared_buy_mw),
            "cleared_sell_mw": _safe_round(p.cleared_sell_mw),
        })
    return rows


async def _fetch_audit_rows(req: ExportRequest, db: AsyncSession) -> list[dict]:
    """
    Audit = historical actuals + forecasts matched on datetime_block.
    fc_lookup keyed by datetime_block (the only reliable shared key).
    """
    start_dt, end_dt = _range_to_dt_range(req.start_date, req.end_date)

    # ── Actuals ───────────────────────────────────────────────
    hist_stmt = (
        select(HistoricalPrice)
        .where(
            and_(
                HistoricalPrice.market == req.market,
                HistoricalPrice.region == req.region,
                HistoricalPrice.datetime_block >= start_dt,
                HistoricalPrice.datetime_block <= end_dt,
            )
        )
        .order_by(HistoricalPrice.datetime_block)
    )
    prices = (await db.execute(hist_stmt)).scalars().all()

    if not prices:
        raise HTTPException(
            status_code=404,
            detail=(
                f"No historical data for {req.market}/{req.region} "
                f"between {req.start_date} and {req.end_date}."
            ),
        )

    # ── Forecasts for same datetime_block range ───────────────
    # Join ForecastRun only for market/region context; filter on
    # Forecast.datetime_block directly (avoids the missing Forecast.forecast_date).
    fc_stmt = (
        select(Forecast)
        .join(ForecastRun, Forecast.forecast_run_id == ForecastRun.id)
        .where(
            and_(
                Forecast.market == req.market,
                Forecast.region == req.region,
                Forecast.datetime_block >= start_dt,
                Forecast.datetime_block <= end_dt,
            )
        )
    )
    # Key: datetime_block → Forecast  (exact match used in audit join)
    fc_lookup: dict[datetime, Forecast] = {
        f.datetime_block: f
        for f in (await db.execute(fc_stmt)).scalars().all()
    }

    rows = []
    for p in prices:
        fc = fc_lookup.get(p.datetime_block)  # match on datetime_block

        actual = float(p.mcp_rs_mwh) if p.mcp_rs_mwh is not None else None
        predicted = float(fc.predicted_price) if fc and fc.predicted_price is not None else None

        abs_err = (
            round(abs(actual - predicted), 2)
            if actual is not None and predicted is not None
            else ""
        )
        pct_err = (
            round(abs(actual - predicted) / actual * 100, 2)
            if actual is not None and predicted is not None and actual != 0
            else ""
        )

        row: dict = {
            "date": str(p.datetime_block.date()),
            "block_no": _dt_to_block_no(p.datetime_block),
            "block_time": _dt_to_time_str(p.datetime_block),
            "actual_mcp_rs_mwh": _safe_round(actual) if actual is not None else "",
            "predicted_mcp_rs_mwh": _safe_round(predicted) if predicted is not None else "",
            "abs_error_rs_mwh": abs_err,
            "pct_error_pct": pct_err,
        }
        if req.include_ci:
            row["lower_ci_p10_rs_mwh"] = _safe_round(fc.lower_ci) if fc else ""
            row["upper_ci_p90_rs_mwh"] = _safe_round(fc.upper_ci) if fc else ""

        rows.append(row)
    return rows


# ══════════════════════════════════════════════════════
#  File Generators
# ══════════════════════════════════════════════════════

_HEADER_LABELS: dict[str, str] = {
    "date": "Date",
    "block_no": "Block",
    "block_time": "Time (IST)",
    "predicted_mcp_rs_mwh": "Predicted Price (Rs/MWh)",
    "lower_ci_p10_rs_mwh": "P10 Lower CI (Rs/MWh)",
    "upper_ci_p90_rs_mwh": "P90 Upper CI (Rs/MWh)",
    "actual_mcp_rs_mwh": "Actual Price (Rs/MWh)",
    "predicted_mcp_rs_mwh": "Predicted Price (Rs/MWh)",
    "cleared_buy_mw": "Cleared Buy (MW)",
    "cleared_sell_mw": "Cleared Sell (MW)",
    "abs_error_rs_mwh": "Abs Error (Rs/MWh)",
    "pct_error_pct": "Error (%)",
}


def _generate_csv(rows: list[dict]) -> str:
    if not rows:
        return "No data available for the selected criteria.\n"
    keys = list(rows[0].keys())
    lines = [",".join(_HEADER_LABELS.get(k, k) for k in keys)]
    for row in rows:
        lines.append(",".join(str(row.get(k, "")) for k in keys))
    return "\n".join(lines) + "\n"


def _generate_xlsx(rows: list[dict], req: ExportRequest) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{req.market} {req.export_type.title()}"[:31]

    C_DARK   = "0d1117"
    C_HDR    = "16213e"
    C_ROW_ALT = "eef2f7"
    C_ERR    = "dc2626"
    C_WARN   = "d97706"

    num_cols = len(rows[0]) if rows else 5

    # ── Banner ────────────────────────────────────────────────
    ws.append(["tatva.gridprice  —  Power Price Intelligence"])
    ws.merge_cells(f"A1:{get_column_letter(num_cols)}1")
    c = ws["A1"]
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Meta ──────────────────────────────────────────────────
    for meta in _build_meta_rows(req):
        ws.append([meta])
        idx = ws.max_row
        ws.merge_cells(f"A{idx}:{get_column_letter(num_cols)}{idx}")
        c = ws.cell(idx, 1)
        c.font = Font(size=9, color="888888", italic=True, name="Calibri")
        c.fill = PatternFill("solid", fgColor=C_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[idx].height = 15

    ws.append([""])           # spacer
    ws.row_dimensions[ws.max_row].height = 6

    if not rows:
        ws.append(["No data found."])
        out = io.BytesIO(); wb.save(out); out.seek(0); return out.read()

    # ── Column headers ────────────────────────────────────────
    keys = list(rows[0].keys())
    ws.append([_HEADER_LABELS.get(k, k) for k in keys])
    hdr_idx = ws.max_row

    for col_i in range(1, len(keys) + 1):
        c = ws.cell(hdr_idx, col_i)
        c.fill = PatternFill("solid", fgColor=C_HDR)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr_idx].height = 36

    # ── Data rows ─────────────────────────────────────────────
    for row_i, row in enumerate(rows):
        ws.append([row.get(k, "") for k in keys])
        data_idx = ws.max_row
        fill = C_ROW_ALT if row_i % 2 == 0 else "ffffff"

        for col_i, key in enumerate(keys, 1):
            c = ws.cell(data_idx, col_i)
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=9, name="Calibri")

            if key == "pct_error_pct":
                val = row.get(key)
                if isinstance(val, (int, float)):
                    if val > 15:
                        c.font = Font(size=9, bold=True, color=C_ERR, name="Calibri")
                    elif val > 8:
                        c.font = Font(size=9, bold=True, color=C_WARN, name="Calibri")

    # ── Column widths ─────────────────────────────────────────
    for col_i, key in enumerate(keys, 1):
        label_len = len(_HEADER_LABELS.get(key, key))
        data_max = max((len(str(r.get(key, ""))) for r in rows), default=0)
        ws.column_dimensions[get_column_letter(col_i)].width = min(
            max(label_len, data_max) + 4, 24
        )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _build_meta_rows(req: ExportRequest) -> list[str]:
    lines = [
        f"Market: {req.market}   |   Region: {req.region}   |   Type: {req.export_type.title()}",
    ]
    if req.export_type == "forecast":
        lines.append(f"Forecast Date: {req.export_date}")
    else:
        lines.append(f"Period: {req.start_date}  →  {req.end_date}")
    lines.append(
        "Confidence Intervals: Included (P10 / P90)"
        if req.include_ci
        else "Confidence Intervals: Excluded"
    )
    lines.append(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  (IST = UTC +5:30)"
    )
    return lines


def _build_filename(req: ExportRequest) -> str:
    if req.export_type == "forecast":
        return f"gridprice_{req.market}_{req.region}_forecast_{req.export_date}"
    return (
        f"gridprice_{req.market}_{req.region}_{req.export_type}"
        f"_{req.start_date}_to_{req.end_date}"
    )


# ══════════════════════════════════════════════════════
#  Audit log  (non-blocking — export always succeeds)
# ══════════════════════════════════════════════════════

async def _log_export(
    db: AsyncSession,
    user_id: uuid.UUID,
    req: ExportRequest,
    file_format: str,
) -> None:
    try:
        job = ExportJob(
            id=uuid.uuid4(),
            user_id=user_id,
            export_type=f"{req.export_type}:{file_format}",
            status="completed",
            file_path=None,
        )
        db.add(job)
        await db.commit()
    except Exception:
        await db.rollback()


# ══════════════════════════════════════════════════════
#  Route Handlers
# ══════════════════════════════════════════════════════

async def _resolve(
    req: ExportRequest, db: AsyncSession
) -> tuple[list[dict], str]:
    if req.export_type == "forecast":
        rows = await _fetch_forecast_rows(req, db)
    elif req.export_type == "historical":
        rows = await _fetch_historical_rows(req, db)
    else:
        rows = await _fetch_audit_rows(req, db)
    return rows, _build_filename(req)


@router.post("/csv")
async def export_csv(
    req: ExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rows, filename = await _resolve(req, db)
    csv_text = _generate_csv(rows)
    await _log_export(db, uuid.UUID(str(current_user.id)), req, "csv")
    return StreamingResponse(
        io.StringIO(csv_text),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


@router.post("/xlsx")
async def export_xlsx(
    req: ExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    rows, filename = await _resolve(req, db)
    xlsx_bytes = _generate_xlsx(rows, req)
    await _log_export(db, uuid.UUID(str(current_user.id)), req, "xlsx")
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )